# -*- coding: utf-8 -*-
"""
更新資料：把同層的 ../股票.xlsx 重新解析寫進 data.js
用法：
  pip install openpyxl
  python rebuild_data.py
之後重新整理 index.html 就會看到新資料。
"""
import os, json, sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("請先安裝 openpyxl：pip install openpyxl")
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.normpath(os.path.join(HERE, '..', '股票.xlsx'))
OUT = os.path.join(HERE, 'data.js')

if not os.path.exists(XLSX):
    print(f"找不到 {XLSX}，請確認檔名跟位置")
    sys.exit(1)

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['股票']

records = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None: continue
    who, year, buy_date, name, code, buy_price, units, buy_total, sell_date, sell_price, sell_units, sell_total, dividend = row[:13]
    # 配股 units 應為整數（券商實際配發都是整股；Excel 公式有時會留小數尾差）
    if who == '配股' and isinstance(units, (int, float)):
        units = round(units)
    records.append({
        'who': who,
        'year': year if isinstance(year, int) else None,
        'buy_date': buy_date.strftime('%Y-%m-%d') if isinstance(buy_date, datetime) else None,
        'name': name,
        'code': str(code) if code is not None else None,
        'buy_price': buy_price,
        'units': units,
        'buy_total': buy_total,
        'sell_date': sell_date.strftime('%Y-%m-%d') if isinstance(sell_date, datetime) else None,
        'sell_price': sell_price,
        'sell_units': sell_units,
        'sell_total': sell_total,
        'dividend': dividend,
    })

# 從 股票總覽 推算最新現價（持股單位 × 現價 = P 欄市值）
# 若 cache 失效讀不到，使用內建 fallback
FALLBACK_PRICES = {
    '2002': 18.75, '00881': 46.86, '00878': 25.30, '2887': 24.40,
    '2881': 88.60, '2882': 74.70, '9945': 23.90, '00713': 53.00,
    '0056': 41.11, '00929': 22.36, '00919': 23.50, '0050': 92.00,
    '00937B': 14.78, '2883': 21.30, '6285': 217.50,
    '2888': 8.33, '3035': 70, '3231': 100,
}
latest = {}
ws2 = wb['股票總覽']
for row in ws2.iter_rows(min_row=4, max_row=20, values_only=True):
    name, units, cost, avg = row[8], row[9], row[10], row[11]
    market = row[15]
    if not name or not units or not market: continue
    # find code by name
    code = None
    for r in records:
        if r['name'] == name and r['code']:
            code = r['code']; break
    if code:
        try: latest[code] = round(market / units, 4)
        except: pass

# 補漏：用最後一次配息那天的價格 (沒有也沒關係)
code_to_name = {}
for r in records:
    if r['code'] and r['name']:
        code_to_name[r['code']] = r['name']

etf_codes = ['0050','0056','00713','00878','00881','00919','00929','00937B']

# fallback：補上沒抓到的
for code, price in FALLBACK_PRICES.items():
    if code not in latest:
        latest[code] = price

data = {
    'records': records,
    'latest_prices': latest,
    'code_to_name': code_to_name,
    'etf_codes': etf_codes,
    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
}

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('window.STOCK_DATA = ')
    json.dump(data, f, ensure_ascii=False)
    f.write(';\n')

print(f"OK: {len(records)} 筆紀錄，{len(code_to_name)} 檔股票")
print(f"已寫入 {OUT}")
print("打開 index.html 並按 Ctrl+Shift+R（強制重新整理）即可看到新資料")
